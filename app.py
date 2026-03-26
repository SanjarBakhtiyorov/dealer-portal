import streamlit as st
import pandas as pd
import numpy as np
import math
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PAGE CONFIG ──────────────────────────────────────────────
st.set_page_config(
    page_title="NormDealer — Нормативная модель",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CUSTOM CSS ───────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=Manrope:wght@400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'Manrope', sans-serif; }

.stApp { background: #0e0f11; color: #f0eeea; }

[data-testid="stSidebar"] {
    background: #161719 !important;
    border-right: 1px solid rgba(255,255,255,0.07);
}
[data-testid="stSidebar"] * { color: #f0eeea !important; }

/* Metric cards */
[data-testid="metric-container"] {
    background: #161719;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 12px;
    padding: 16px 20px;
}
[data-testid="stMetricLabel"] { color: #7a7975 !important; font-size: 11px !important; text-transform: uppercase; letter-spacing: .08em; font-family: 'IBM Plex Mono', monospace !important; }
[data-testid="stMetricValue"] { color: #f0eeea !important; font-family: 'Manrope', sans-serif !important; font-weight: 700 !important; }
[data-testid="stMetricDelta"] { font-family: 'IBM Plex Mono', monospace !important; font-size: 11px !important; }

/* Tabs */
[data-testid="stTabs"] button {
    font-family: 'Manrope', sans-serif !important;
    font-weight: 500;
    color: #7a7975 !important;
    border-radius: 8px;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #c8f04a !important;
    border-bottom-color: #c8f04a !important;
}

/* Buttons */
.stButton > button {
    background: #1e2024;
    border: 1px solid rgba(255,255,255,0.14);
    color: #f0eeea;
    border-radius: 8px;
    font-family: 'Manrope', sans-serif;
    font-weight: 500;
    transition: all .15s;
}
.stButton > button:hover {
    border-color: rgba(200,240,74,0.4);
    color: #c8f04a;
}

/* Accent button */
.accent-btn > button {
    background: #c8f04a !important;
    color: #0e0f11 !important;
    border: none !important;
    font-weight: 600 !important;
}
.accent-btn > button:hover { background: #d8ff5a !important; }

/* Selectbox, inputs */
.stSelectbox > div, .stNumberInput > div input, .stTextInput > div input {
    background: #1e2024 !important;
    border: 1px solid rgba(255,255,255,0.14) !important;
    color: #f0eeea !important;
    border-radius: 8px !important;
    font-family: 'Manrope', sans-serif !important;
}

/* Dataframe */
[data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: #161719;
    border: 1.5px dashed rgba(255,255,255,0.2);
    border-radius: 10px;
    padding: 8px;
}
[data-testid="stFileUploader"]:hover { border-color: rgba(200,240,74,0.4); }

/* Expander */
[data-testid="stExpander"] {
    background: #161719 !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 10px !important;
}

/* Success / warning / info alerts */
[data-testid="stAlert"] { border-radius: 8px !important; }

/* Header override */
h1, h2, h3 { color: #f0eeea !important; font-family: 'Manrope', sans-serif !important; }

/* Divider */
hr { border-color: rgba(255,255,255,0.07) !important; }

/* Mono text helper */
.mono { font-family: 'IBM Plex Mono', monospace; font-size: 12px; color: #c8f04a; }
.muted { color: #7a7975; font-size: 12px; }

/* Pill badges */
.pill-ok  { background:#162012; color:#c8f04a; padding:2px 10px; border-radius:20px; font-size:11px; font-family:'IBM Plex Mono',monospace; }
.pill-warn{ background:#1f1800; color:#f0a832; padding:2px 10px; border-radius:20px; font-size:11px; font-family:'IBM Plex Mono',monospace; }
.pill-bad { background:#200a08; color:#e85c3e; padding:2px 10px; border-radius:20px; font-size:11px; font-family:'IBM Plex Mono',monospace; }
</style>
""", unsafe_allow_html=True)


# ── DEALER DATA ──────────────────────────────────────────────
DEALERS = {
    "Бахтиёр Абу-Сахий":    {"sap": "1000001081", "rev": 1303952, "pallets": 11101, "area": 5002.8, "gross_pct": 0.0592},
    "Улугбек Самарканд":     {"sap": "1000000029", "rev": 1060885, "pallets": 4854,  "area": 2187.5, "gross_pct": 0.0607},
    "Шодмон Денов":          {"sap": "1000000430", "rev": 1040169, "pallets": 5763,  "area": 2597.2, "gross_pct": 0.0635},
    "Ахмаджон Коканд":       {"sap": "1000000646", "rev": 721601,  "pallets": 8070,  "area": 3636.8, "gross_pct": 0.0658},
    "Шавкат Самарканд":      {"sap": "1000000766", "rev": 759541,  "pallets": 4623,  "area": 2083.4, "gross_pct": 0.0592},
    "Аскар Карши":           {"sap": "1000000234", "rev": 615169,  "pallets": 5190,  "area": 2338.9, "gross_pct": 0.0558},
    "Шариф Термез":          {"sap": "1000000426", "rev": 697567,  "pallets": 6327,  "area": 2851.3, "gross_pct": 0.0503},
    "Акмал Бухара":          {"sap": "1000001233", "rev": 497470,  "pallets": 6131,  "area": 2763.0, "gross_pct": 0.0594},
    "Обид Бухара":           {"sap": "1000000564", "rev": 511097,  "pallets": 3269,  "area": 1473.2, "gross_pct": 0.0599},
    "Кахрамон Навои":        {"sap": "1000001100", "rev": 375131,  "pallets": 7805,  "area": 3517.4, "gross_pct": 0.0506},
}


# ── CALCULATION ENGINE ───────────────────────────────────────
def calc_norms(rev, pallets, area, gross_pct, assumptions):
    a = assumptions
    gross = rev * gross_pct

    # Personnel
    mgr_count = max(1, math.ceil(rev / a["rev_per_mgr"]))
    log_count = max(1, math.floor(pallets / 5))
    personnel = (
        mgr_count * a["sal_mgr"] +
        a["sal_acc"] +
        log_count * a["sal_log"] +
        a["sal_off"] +
        a["sal_dir"]
    )

    # Taxes
    tax_vat     = gross * a["vat"]
    tax_profit  = gross * a["profit_tax"]
    tax_payroll = personnel * a["payroll_tax"]
    taxes_total = tax_vat + tax_profit + tax_payroll

    # Warehouse
    norm_area   = area / a["floors"]
    rent        = norm_area * a["rent_per_m2"]
    trips       = max(1, math.ceil(pallets / 500) * 2)
    loading     = trips * a["trip_cost"]
    warehouse_total = rent + loading

    # Other
    fuel    = 200 + trips * 40
    mobile  = 100 + mgr_count * 30
    internet= 50
    banking = rev * 0.001
    other_total = fuel + mobile + internet + banking

    total = personnel + taxes_total + warehouse_total + other_total
    net_profit = gross - total

    return {
        "rev": rev, "gross": gross,
        "personnel": personnel, "mgr_count": mgr_count, "log_count": log_count,
        "tax_vat": tax_vat, "tax_profit": tax_profit, "tax_payroll": tax_payroll, "taxes": taxes_total,
        "norm_area": norm_area, "rent": rent, "trips": trips, "loading": loading, "warehouse": warehouse_total,
        "fuel": fuel, "mobile": mobile, "internet": internet, "banking": banking, "other": other_total,
        "total": total, "net_profit": net_profit, "pct_of_rev": total / rev if rev else 0,
    }


def process_uploaded_stock(df):
    """Parse uploaded Excel and compute pallets & area."""
    df.columns = [str(c).strip() for c in df.columns]
    # Try to find relevant columns flexibly
    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if any(x in cl for x in ["артикул", "article", "sku", "код"]):
            col_map["sku"] = c
        elif any(x in cl for x in ["остаток", "кол-во", "qty", "quantity", "stock"]):
            col_map["qty"] = c
        elif any(x in cl for x in ["шт/поддон", "поддон", "pallet", "шт/п"]):
            col_map["ppu"] = c
        elif any(x in cl for x in ["объем", "объём", "volume", "м³"]):
            col_map["vol"] = c

    if "qty" not in col_map or "ppu" not in col_map:
        return None, "Не найдены обязательные колонки: «Остаток (шт)» и «Шт/поддон»"

    df = df.rename(columns=col_map)
    df["qty"] = pd.to_numeric(df.get("qty", 0), errors="coerce").fillna(0)
    df["ppu"] = pd.to_numeric(df.get("ppu", 1), errors="coerce").fillna(1).replace(0, 1)

    df["pallets"] = np.ceil(df["qty"] / df["ppu"])
    total_pallets = int(df["pallets"].sum())

    # Area calc
    if "vol" in col_map:
        df["vol"] = pd.to_numeric(df["vol"], errors="coerce").fillna(0.42)
        avg_vol = df.loc[df["vol"] > 0, "vol"].mean() or 0.42
    else:
        avg_vol = 0.42  # default m³

    pallet_footprint = 0.322  # m²
    avg_pallet_h = 0.75       # m
    rack_height = 6.0
    tiers = math.floor(rack_height / avg_pallet_h)
    pallet_places_with_tiers = math.ceil(total_pallets / tiers)
    area = pallet_places_with_tiers * pallet_footprint * 1.4  # aisle factor

    return {"pallets": total_pallets, "area": round(area, 1), "sku_count": len(df[df["qty"] > 0]), "avg_vol": round(avg_vol, 3)}, None


# ── DEFAULT ASSUMPTIONS ──────────────────────────────────────
DEFAULT_ASSUMPTIONS = {
    "vat": 0.12, "profit_tax": 0.15, "payroll_tax": 0.12,
    "sal_mgr": 600, "sal_acc": 700, "sal_log": 400, "sal_off": 500, "sal_dir": 1200,
    "rev_per_mgr": 350000, "rent_per_m2": 2.0, "trip_cost": 50, "floors": 2,
}

if "assumptions" not in st.session_state:
    st.session_state.assumptions = DEFAULT_ASSUMPTIONS.copy()
if "stock_override" not in st.session_state:
    st.session_state.stock_override = {}  # dealer_name -> {pallets, area}


# ── HELPERS ─────────────────────────────────────────────────
def fmt(n): return f"${n:,.0f}"
def fmtpct(n): return f"{n*100:.2f}%"


def get_dealer_data(name):
    d = DEALERS[name].copy()
    if name in st.session_state.stock_override:
        d.update(st.session_state.stock_override[name])
    return d


def build_excel_report(name, n, d):
    wb = Workbook()
    ws = wb.active
    ws.title = "Норматив"
    thin = Side(style="thin", color="2A2D32")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, val, bold=False, bg="1E2024", fg="C8F04A"):
        c = ws[cell]; c.value = val
        c.font = Font(name="Calibri", bold=bold, size=11, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = brd
        c.alignment = Alignment(horizontal="left", indent=1, vertical="center")

    def row(r, label, val, pct=None):
        ws[f"A{r}"].value = label
        ws[f"A{r}"].font = Font(name="Calibri", size=10, color="F0EEEA")
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="0E1011")
        ws[f"A{r}"].border = brd
        ws[f"B{r}"].value = val
        ws[f"B{r}"].font = Font(name="Calibri", size=10, color="C8F04A")
        ws[f"B{r}"].fill = PatternFill("solid", fgColor="0E1011")
        ws[f"B{r}"].border = brd
        ws[f"B{r}"].number_format = '$#,##0'
        if pct is not None:
            ws[f"C{r}"].value = pct
            ws[f"C{r}"].font = Font(name="Calibri", size=10, color="7A7975")
            ws[f"C{r}"].fill = PatternFill("solid", fgColor="0E1011")
            ws[f"C{r}"].border = brd
            ws[f"C{r}"].number_format = '0.00%'

    ws.merge_cells("A1:C1")
    ws["A1"].value = f"Нормативная модель — {name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="C8F04A")
    ws["A1"].fill = PatternFill("solid", fgColor="1E2024")
    ws["A1"].alignment = Alignment(horizontal="left", indent=1, vertical="center")
    ws.row_dimensions[1].height = 28

    hdr("A2", "Статья", bold=True); hdr("B2", "$/мес", bold=True); hdr("C2", "% выручки", bold=True)

    data = [
        ("Выручка",              n["rev"],       None),
        ("Валовая прибыль",      n["gross"],      n["gross"] / n["rev"]),
        ("",                     None,            None),
        ("Персонал",             n["personnel"],  n["personnel"] / n["rev"]),
        ("Налоги",               n["taxes"],      n["taxes"] / n["rev"]),
        ("Склад и логистика",    n["warehouse"],  n["warehouse"] / n["rev"]),
        ("Прочие расходы",       n["other"],      n["other"] / n["rev"]),
        ("ИТОГО РАСХОДЫ",        n["total"],      n["pct_of_rev"]),
        ("",                     None,            None),
        ("Чистая прибыль",       n["net_profit"], n["net_profit"] / n["rev"]),
    ]
    for i, (label, val, pct) in enumerate(data, start=3):
        row(i, label, val, pct)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_stock_template():
    """Generate the blank upload template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Сток дилера"
    thin = Side(style="thin", color="2A2D32")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADER_BG, HEADER_FG = "1E2024", "C8F04A"
    REQ_BG, OPT_BG = "162012", "0E1A2A"
    DIM = "7A7975"; WHITE = "F0EEEA"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "ШАБЛОН ЗАГРУЗКИ СТОКА — Нормативная модель дилера"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=HEADER_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"].value = "Заполните колонки A, C, D. Колонки B, E, F, G — необязательны."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=DIM)
    ws["A2"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    headers = [
        ("A", "Артикул *",       REQ_BG, HEADER_FG),
        ("B", "Описание",        OPT_BG, DIM),
        ("C", "Остаток (шт) *",  REQ_BG, HEADER_FG),
        ("D", "Шт/поддон *",     REQ_BG, HEADER_FG),
        ("E", "Объем ед. (м³)",  OPT_BG, DIM),
        ("F", "Ширина (см)",     OPT_BG, DIM),
        ("G", "Высота (см)",     OPT_BG, DIM),
    ]
    for col_letter, title, bg, fg in headers:
        c = ws[f"{col_letter}4"]
        c.value = title
        c.font = Font(name="Calibri", bold=True, size=11, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    ws.row_dimensions[4].height = 28

    examples = [
        ("FKN90014BEL-IN",   "Конд ВБ ARCSIJ3AW09BE Crystal Inv.",  40, 1, 0.423, 65, 93),
        ("FKN12021BEL/S-IN", "Конд ВБ SHVSIJ2AW12BE Sensei Inv.",  540, 1, 0.423, 65, 93),
        ("FKN12044BEL/A-IN", "Конд ВБ AVLSIQ1AW12BE Brugge Inv.", 3004, 1, 0.423, 65, 93),
        ("FTHD001076SER/D",  "Холод Samsung RB29 FERNDSA Серый",    87, 1, 0.380, 60, 167),
        ("FKV0760BEL",       "Кух вытяжка ART-0760 Белый Prima",    10, 2, 0.212, 60, 45),
    ]
    for i, row_data in enumerate(examples):
        r = i + 5
        for j, val in enumerate(row_data):
            col = get_column_letter(j + 1)
            c = ws[f"{col}{r}"]
            c.value = val
            c.fill = PatternFill("solid", fgColor="141A10" if j in [0,2,3] else "111316")
            c.font = Font(name="Calibri", size=10, color=WHITE if j in [0,2,3] else DIM)
            c.alignment = Alignment(horizontal="left" if j < 2 else "center", vertical="center", indent=1)
            c.border = brd
        ws.row_dimensions[r].height = 20

    for r in range(10, 60):
        for j in range(7):
            col = get_column_letter(j + 1)
            c = ws[f"{col}{r}"]
            c.fill = PatternFill("solid", fgColor="0E1011")
            c.border = brd
            c.font = Font(name="Calibri", size=10, color=WHITE)
            c.alignment = Alignment(horizontal="left" if j < 2 else "center", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════
#  SIDEBAR
# ════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 📊 NormDealer")
    st.markdown("<div class='muted' style='font-size:11px;margin-bottom:16px'>Нормативная модель · 2026</div>", unsafe_allow_html=True)

    dealer_name = st.selectbox("Выберите дилера", list(DEALERS.keys()), label_visibility="collapsed")
    st.divider()

    st.markdown("**⚙️ Допущения модели**")

    with st.expander("💰 Оклады ($/мес)", expanded=False):
        st.session_state.assumptions["sal_mgr"] = st.number_input("Менеджер по продажам", value=st.session_state.assumptions["sal_mgr"], step=50)
        st.session_state.assumptions["sal_acc"] = st.number_input("Бухгалтер",            value=st.session_state.assumptions["sal_acc"], step=50)
        st.session_state.assumptions["sal_log"] = st.number_input("Логистика",            value=st.session_state.assumptions["sal_log"], step=50)
        st.session_state.assumptions["sal_off"] = st.number_input("Офис менеджер",        value=st.session_state.assumptions["sal_off"], step=50)
        st.session_state.assumptions["sal_dir"] = st.number_input("Директор",             value=st.session_state.assumptions["sal_dir"], step=50)

    with st.expander("🏦 Налоги", expanded=False):
        st.session_state.assumptions["vat"]         = st.number_input("НДС (%)",              value=st.session_state.assumptions["vat"]*100,        step=1.0) / 100
        st.session_state.assumptions["profit_tax"]  = st.number_input("Налог на прибыль (%)", value=st.session_state.assumptions["profit_tax"]*100,  step=1.0) / 100
        st.session_state.assumptions["payroll_tax"] = st.number_input("Зарплатный налог (%)", value=st.session_state.assumptions["payroll_tax"]*100, step=1.0) / 100

    with st.expander("🏭 Склад", expanded=False):
        st.session_state.assumptions["rent_per_m2"] = st.number_input("Аренда ($/м²/мес)", value=st.session_state.assumptions["rent_per_m2"], step=0.5)
        st.session_state.assumptions["floors"]      = st.number_input("Этажность склада",  value=st.session_state.assumptions["floors"],      step=1, min_value=1)
        st.session_state.assumptions["trip_cost"]   = st.number_input("Стоимость рейса ($)",value=st.session_state.assumptions["trip_cost"],  step=10)

    st.divider()
    st.markdown("<div class='mono'>● Сток синхронизирован</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
#  MAIN CONTENT
# ════════════════════════════════════════════════════════════
d = get_dealer_data(dealer_name)
n = calc_norms(d["rev"], d["pallets"], d["area"], d["gross_pct"], st.session_state.assumptions)

# Header
col_h1, col_h2 = st.columns([3, 1])
with col_h1:
    st.markdown(f"## {dealer_name}")
    st.markdown(f"<div class='muted'>SAP: {d['sap']} · Паллетомест: {d['pallets']:,} · Площадь: {d['area']:,.0f} м²</div>", unsafe_allow_html=True)
with col_h2:
    report_buf = build_excel_report(dealer_name, n, d)
    st.download_button("⬇ Скачать отчёт Excel", data=report_buf, file_name=f"норматив_{dealer_name.replace(' ','_')}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

st.divider()

# ── TABS ────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["📈 Дашборд", "📋 Детализация расходов", "📂 Загрузить данные стока", "⊞ Все дилеры"])


# ════ TAB 1: DASHBOARD ══════════════════════════════════════
with tab1:
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.metric("Выручка / месяц",         fmt(d["rev"]),        delta="активный дилер")
    with c2: st.metric("Валовая прибыль",          fmt(n["gross"]),      delta=fmtpct(d["gross_pct"]) + " маржа")
    with c3: st.metric("Норм. расходы",            fmt(n["total"]),      delta=fmtpct(n["pct_of_rev"]) + " от выручки", delta_color="inverse")
    with c4:
        delta_label = "✓ Дилер жизнеспособен" if n["net_profit"] > 0 else "⚠ Проверить расходы"
        st.metric("Чистая прибыль",                fmt(n["net_profit"]), delta=delta_label)

    st.markdown("---")
    left_col, right_col = st.columns([3, 2])

    with left_col:
        st.markdown("**Нормативные расходы по категориям**")

        categories = [
            ("👤 Персонал",          n["personnel"], d["rev"], "#5b9ef5"),
            ("🏦 Налоги",            n["taxes"],     d["rev"], "#f0a832"),
            ("🏭 Склад и логистика", n["warehouse"], d["rev"], "#e85c3e"),
            ("⚡ Прочие расходы",    n["other"],     d["rev"], "#c8f04a"),
        ]

        df_cats = pd.DataFrame({
            "Категория": [c[0] for c in categories],
            "$/мес":     [c[1] for c in categories],
            "% выручки": [c[1]/c[2]*100 for c in categories],
        })

        max_val = max(c[1] for c in categories)
        for cat, val, rev, color in categories:
            bar_pct = int(val / max_val * 100)
            pct_str = f"{val/rev*100:.2f}%"
            st.markdown(f"""
            <div style="background:#161719;border:1px solid rgba(255,255,255,0.07);border-radius:10px;
                        padding:14px 16px;margin-bottom:8px;display:flex;align-items:center;gap:14px">
                <span style="font-size:18px;width:24px">{cat.split()[0]}</span>
                <div style="flex:1">
                    <div style="font-size:13px;font-weight:600;margin-bottom:6px">{' '.join(cat.split()[1:])}</div>
                    <div style="height:4px;background:rgba(255,255,255,0.07);border-radius:2px">
                        <div style="height:4px;width:{bar_pct}%;background:{color};border-radius:2px"></div>
                    </div>
                </div>
                <div style="text-align:right">
                    <div style="font-family:'IBM Plex Mono',monospace;font-size:13px;font-weight:600">{fmt(val)}</div>
                    <div style="font-size:11px;color:#7a7975">{pct_str}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

    with right_col:
        st.markdown("**Статус нормативов**")

        if n["pct_of_rev"] < 0.025:
            st.success(f"✓ Расходы {fmtpct(n['pct_of_rev'])} от выручки — в норме (цель < 2.5%)")
        elif n["pct_of_rev"] < 0.04:
            st.warning(f"⚠ Расходы {fmtpct(n['pct_of_rev'])} — немного выше ориентира")
        else:
            st.error(f"✗ Расходы {fmtpct(n['pct_of_rev'])} — выше нормы, требует внимания")

        if n["net_profit"] > 0:
            st.success(f"✓ Чистая прибыль: {fmt(n['net_profit'])}/мес")
        else:
            st.error(f"✗ Убыток: {fmt(n['net_profit'])}/мес")

        st.info(f"ℹ Нормативная площадь склада: {n['norm_area']:.0f} м²")
        st.info(f"ℹ Рейсов в месяц: {n['trips']}")

        st.markdown("---")
        st.markdown("**Склад & сток**")
        stock_data = {
            "Паллетомест":         f"{d['pallets']:,}",
            "Полная площадь (м²)": f"{d['area']:,.0f}",
            "Норм. площадь (м²)":  f"{n['norm_area']:.0f}",
            "Этажность":           str(st.session_state.assumptions["floors"]),
            "Аренда / мес":        fmt(n["rent"]),
        }
        for k, v in stock_data.items():
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;padding:7px 0;
                        border-bottom:1px solid rgba(255,255,255,0.05);font-size:12px">
                <span style="color:#7a7975">{k}</span>
                <span style="font-family:'IBM Plex Mono',monospace;color:#c8f04a">{v}</span>
            </div>""", unsafe_allow_html=True)

    # Comparison table
    st.markdown("---")
    st.markdown("**Структура затрат vs. норматив**")
    bench_data = [
        {"Категория": "Персонал",          "Сумма, $/мес": n["personnel"], "% выручки": n["personnel"]/d["rev"], "Ориентир": "0.8–1.4%", "Норм. сумма": d["rev"]*0.011},
        {"Категория": "Налоги",            "Сумма, $/мес": n["taxes"],     "% выручки": n["taxes"]/d["rev"],     "Ориентир": "0.4–0.8%", "Норм. сумма": d["rev"]*0.006},
        {"Категория": "Склад и логистика", "Сумма, $/мес": n["warehouse"], "% выручки": n["warehouse"]/d["rev"], "Ориентир": "0.5–0.9%", "Норм. сумма": d["rev"]*0.007},
        {"Категория": "Прочее",            "Сумма, $/мес": n["other"],     "% выручки": n["other"]/d["rev"],     "Ориентир": "0.3–0.7%", "Норм. сумма": d["rev"]*0.005},
        {"Категория": "ИТОГО",             "Сумма, $/мес": n["total"],     "% выручки": n["pct_of_rev"],         "Ориентир": "2.0–3.5%", "Норм. сумма": d["rev"]*0.035},
    ]
    df_bench = pd.DataFrame(bench_data)
    df_bench["Статус"] = df_bench.apply(
        lambda r: "✓ В норме" if r["% выручки"] <= r["Норм. сумма"]/d["rev"]*1.2 else "⚠ Выше нормы", axis=1
    )
    df_bench["Сумма, $/мес"] = df_bench["Сумма, $/мес"].apply(lambda x: f"${x:,.0f}")
    df_bench["% выручки"]    = df_bench["% выручки"].apply(lambda x: f"{x*100:.2f}%")
    df_show = df_bench[["Категория", "Сумма, $/мес", "% выручки", "Ориентир", "Статус"]]
    st.dataframe(df_show, use_container_width=True, hide_index=True)


# ════ TAB 2: BREAKDOWN ══════════════════════════════════════
with tab2:
    st.markdown("### Детализация по статьям расходов")

    sections = [
        {
            "title": "1. Персонал",
            "color": "#5b9ef5",
            "rows": [
                (f"Менеджеры по продажам ({n['mgr_count']} чел × ${st.session_state.assumptions['sal_mgr']})", n["mgr_count"] * st.session_state.assumptions["sal_mgr"]),
                (f"Бухгалтер (1 чел × ${st.session_state.assumptions['sal_acc']})",                            st.session_state.assumptions["sal_acc"]),
                (f"Логистика ({n['log_count']} чел × ${st.session_state.assumptions['sal_log']})",             n["log_count"] * st.session_state.assumptions["sal_log"]),
                (f"Офис менеджер (1 чел × ${st.session_state.assumptions['sal_off']})",                        st.session_state.assumptions["sal_off"]),
                (f"Директор (1 чел × ${st.session_state.assumptions['sal_dir']})",                             st.session_state.assumptions["sal_dir"]),
            ],
            "total": n["personnel"],
        },
        {
            "title": "2. Налоги",
            "color": "#f0a832",
            "rows": [
                (f"НДС с валовой прибыли ({fmt(n['gross'])} × {st.session_state.assumptions['vat']*100:.0f}%)",         n["tax_vat"]),
                (f"Налог на прибыль ({fmt(n['gross'])} × {st.session_state.assumptions['profit_tax']*100:.0f}%)",        n["tax_profit"]),
                (f"Зарплатный налог ({fmt(n['personnel'])} × {st.session_state.assumptions['payroll_tax']*100:.0f}%)",   n["tax_payroll"]),
            ],
            "total": n["taxes"],
        },
        {
            "title": "3. Склад и логистика",
            "color": "#e85c3e",
            "rows": [
                (f"Аренда склада ({n['norm_area']:.0f} м² × ${st.session_state.assumptions['rent_per_m2']}/м²)",         n["rent"]),
                (f"Погрузка/разгрузка ({n['trips']} рейсов × ${st.session_state.assumptions['trip_cost']})",             n["loading"]),
            ],
            "total": n["warehouse"],
        },
        {
            "title": "4. Прочие расходы",
            "color": "#c8f04a",
            "rows": [
                ("Топливо (базовая ставка + рейсы)",           n["fuel"]),
                (f"Мобильная связь ({n['mgr_count']+1} сотр.)", n["mobile"]),
                ("Интернет (фиксированно)",                    n["internet"]),
                (f"Банковские комиссии ({fmt(d['rev'])} × 0.1%)", n["banking"]),
            ],
            "total": n["other"],
        },
    ]

    for sec in sections:
        with st.expander(f"{sec['title']}  —  {fmt(sec['total'])}/мес  ({fmtpct(sec['total']/d['rev'])} выручки)", expanded=True):
            rows = []
            for label, val in sec["rows"]:
                rows.append({"Статья": label, "$/мес": f"${val:,.0f}", "% выручки": f"{val/d['rev']*100:.3f}%"})
            rows.append({"Статья": "── Итого", "$/мес": f"${sec['total']:,.0f}", "% выручки": f"{sec['total']/d['rev']*100:.3f}%"})
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ════ TAB 3: UPLOAD ═════════════════════════════════════════
with tab3:
    st.markdown("### Загрузить данные стока")

    col_up1, col_up2 = st.columns(2)

    with col_up1:
        st.markdown("**Загрузка Excel файла**")
        uploaded_file = st.file_uploader(
            "Перетащите .xlsx файл или нажмите Browse",
            type=["xlsx", "xls"],
            label_visibility="collapsed",
        )

        if uploaded_file:
            try:
                df_raw = pd.read_excel(uploaded_file, header=None)
                # Find header row (contains "артикул" or "остаток")
                header_row = 0
                for i, row in df_raw.iterrows():
                    row_str = " ".join(str(v).lower() for v in row)
                    if any(x in row_str for x in ["артикул", "остаток", "sku"]):
                        header_row = i
                        break
                df_data = pd.read_excel(uploaded_file, header=header_row)
                result, err = process_uploaded_stock(df_data)

                if err:
                    st.error(f"❌ {err}")
                else:
                    st.success(f"✓ Файл обработан: {result['sku_count']} SKU с остатком")
                    col_r1, col_r2 = st.columns(2)
                    with col_r1:
                        st.metric("Паллетомест", f"{result['pallets']:,}")
                    with col_r2:
                        st.metric("Расч. площадь склада", f"{result['area']:,.0f} м²")

                    if st.button("✓ Применить к нормативной модели", type="primary"):
                        st.session_state.stock_override[dealer_name] = {
                            "pallets": result["pallets"],
                            "area":    result["area"],
                        }
                        st.success("Данные обновлены. Норматив пересчитан.")
                        st.rerun()
            except Exception as e:
                st.error(f"❌ Ошибка чтения файла: {e}")

        st.markdown("---")
        template_buf = build_stock_template()
        st.download_button(
            "⬇ Скачать шаблон для заполнения",
            data=template_buf,
            file_name="шаблон_сток_дилера.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_up2:
        st.markdown("**Требования к файлу**")
        req_df = pd.DataFrame([
            {"Колонка": "Артикул",        "Тип": "текст",   "Пример": "FKN90014BEL-IN",  "Обязательно": "✓ Да"},
            {"Колонка": "Остаток (шт)",   "Тип": "число",   "Пример": "42",               "Обязательно": "✓ Да"},
            {"Колонка": "Шт/поддон",      "Тип": "число",   "Пример": "1",                "Обязательно": "✓ Да"},
            {"Колонка": "Описание",       "Тип": "текст",   "Пример": "Конд ВБ...",        "Обязательно": "Нет"},
            {"Колонка": "Объем ед. (м³)", "Тип": "число",   "Пример": "0.423",            "Обязательно": "Нет"},
        ])
        st.dataframe(req_df, use_container_width=True, hide_index=True)

        st.info("💡 Названия колонок могут быть на русском или английском. Система определяет их автоматически.")

        if dealer_name in st.session_state.stock_override:
            ovr = st.session_state.stock_override[dealer_name]
            st.success(f"✓ Загружены данные: {ovr['pallets']:,} паллетомест, {ovr['area']:,.0f} м²")
            if st.button("✕ Сбросить к исходным данным"):
                del st.session_state.stock_override[dealer_name]
                st.rerun()


# ════ TAB 4: ALL DEALERS ════════════════════════════════════
with tab4:
    st.markdown("### Все дилеры — сводная таблица")

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Дилеров в системе", len(DEALERS))
    with c2: st.metric("Средняя выручка",   fmt(sum(d["rev"] for d in DEALERS.values()) / len(DEALERS)))

    all_norms = []
    for name, dealer in DEALERS.items():
        d_ = get_dealer_data(name)
        n_ = calc_norms(d_["rev"], d_["pallets"], d_["area"], d_["gross_pct"], st.session_state.assumptions)
        all_norms.append(n_["pct_of_rev"])

    with c3: st.metric("Средний % расходов", fmtpct(sum(all_norms)/len(all_norms)))

    st.markdown("---")
    rows_admin = []
    for name, dealer in DEALERS.items():
        d_ = get_dealer_data(name)
        n_ = calc_norms(d_["rev"], d_["pallets"], d_["area"], d_["gross_pct"], st.session_state.assumptions)
        rows_admin.append({
            "Дилер":            name,
            "Выручка/мес":      f"${d_['rev']:,.0f}",
            "Норм. расходы":    f"${n_['total']:,.0f}",
            "% выручки":        f"{n_['pct_of_rev']*100:.2f}%",
            "Паллетомест":      f"{d_['pallets']:,}",
            "Площадь склада":   f"{d_['area']:,.0f} м²",
            "Чистая прибыль":   f"${n_['net_profit']:,.0f}",
            "Статус":           "✓ Жизнеспособен" if n_["net_profit"] > 0 else "⚠ Убыток",
        })

    df_admin = pd.DataFrame(rows_admin)
    st.dataframe(df_admin, use_container_width=True, hide_index=True)

    st.markdown("---")
    # Export all
    all_buf = io.BytesIO()
    with pd.ExcelWriter(all_buf, engine="openpyxl") as writer:
        df_admin.to_excel(writer, index=False, sheet_name="Все дилеры")
    all_buf.seek(0)
    st.download_button(
        "⬇ Экспорт всех дилеров в Excel",
        data=all_buf,
        file_name="все_дилеры_нормативы.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
